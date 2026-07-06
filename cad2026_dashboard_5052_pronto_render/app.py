from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs
from datetime import datetime, date
import json, os, re, unicodedata
from collections import defaultdict
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.join(BASE_DIR, 'data', 'cad2026.xlsx')
SHEET = 'CAD 2026'


def norm_text(value):
    if value is None:
        return ''
    s = str(value).strip().upper()
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s = re.sub(r'\s+', ' ', s)
    return s


def to_float(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace('%', '')
    if not s:
        return None
    # aceita 1, 1.0, 100, 100%, 0,5 etc.
    s = s.replace('.', '').replace(',', '.') if ',' in s else s
    try:
        n = float(s)
        # Se veio como 100 em vez de 1, converte percentual para fração
        if n > 1 and n <= 100 and ('%' in str(value) or n in (50, 100) or 'EXEC' not in str(value)):
            return n / 100
        return n
    except ValueError:
        return None


def money_float(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace('R$', '').replace(' ', '')
    if not s:
        return 0.0
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == '':
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    return None


def iso_date(value):
    d = date_value(value)
    return d.isoformat() if d else ''


def br_date_label(value):
    d = date_value(value)
    return d.strftime('%d/%m') if d else ''


def motivo_from_obs(obs):
    t = norm_text(obs)
    if not t:
        return 'Sem motivo na observação'
    rules = [
        ('Condições climáticas', r'CHUVA|CLIMATIC|TEMPO|TEMPORAL|VENTO'),
        ('Atraso na primeira obra', r'ATRASO.*PRIMEIRA|PRIMEIRA OBRA'),
        ('Necessita TES', r'NEC\.?\s*TES|\bTES\b'),
        ('Material', r'CABO|MATERIAL|FALTA MATERIAL'),
        ('Risco Operativo', r'CRUZETA|PODRE|POSTE|ESTRUTURA|RISCO OPERATIVO'),
        ('Pedir desligamento', r'DESLIGAMENTO|DESLIGAR|PEDIR DESL'),
        ('Acesso impedido', r'ACESSO|CLIENTE AUSENTE|SEM ACESSO|PORTAO'),
        ('Equipe deslocada / apoio', r'APOIO|DESLOC|EMERGENCIA|PRIORIDADE'),
        ('Cancelada sem detalhe', r'CANCELAD'),
    ]
    for label, pattern in rules:
        if re.search(pattern, t):
            return label
    first = re.split(r'[;,.\n-]+', str(obs).strip())[0][:60].strip()
    return first.capitalize() if first else 'Outros'


def load_rows(path=DEFAULT_FILE):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: row_cells[i] if i < len(row_cells) else None for i in range(len(headers))}
        if row.get('Nº NOTA') is None or str(row.get('Nº NOTA')).strip() == '':
            continue
        aux = norm_text(row.get('AUX'))
        # Regra global: linhas marcadas como REPETIU não entram em nenhum card, tabela ou gráfico.
        if aux == 'REPETIU' or 'REPETIU' in aux:
            continue

        row['DATA_OBJ'] = date_value(row.get('DATA'))
        row['DATA_ISO'] = iso_date(row.get('DATA'))
        row['DATA_LABEL'] = br_date_label(row.get('DATA'))
        row['US PLANEJADO_NUM'] = money_float(row.get('US PLANEJADO'))
        row['US EXECUTADO_NUM'] = money_float(row.get('US EXECUTADO'))
        row['VALOR PROG_NUM'] = money_float(row.get('VALOR PROG'))
        row['VALOR EXEC_NUM'] = money_float(row.get('VALOR  EXEC'))
        row['PERC_EXEC_NUM'] = to_float(row.get('% EXEC'))
        row['EXEC_APONTADA'] = row['PERC_EXEC_NUM'] is not None
        if row['EXEC_APONTADA']:
            row['PERDA_US'] = max(row['US PLANEJADO_NUM'] - row['US EXECUTADO_NUM'], 0)
            row['PERDA_VALOR'] = max(row['VALOR PROG_NUM'] - row['VALOR EXEC_NUM'], 0)
            row['MOTIVO_PERDA'] = motivo_from_obs(row.get('OBS.')) if row['PERC_EXEC_NUM'] < 1 else 'Sem perda'
        rows.append(row)
    return rows



def nota_key(row):
    return str(row.get('Nº NOTA') or '').strip()


def dedupe_by_nota(rows):
    """Mantém somente uma linha por Nº NOTA para não estourar gráficos/KPIs por AUX repetido."""
    best = {}
    for r in rows:
        key = nota_key(r)
        if not key:
            continue
        cur = best.get(key)
        if cur is None:
            best[key] = r
            continue
        # Preferir a linha com maior % EXEC; se empatar, a linha com maior valor executado; se empatar, mantém a primeira.
        r_exec = r.get('PERC_EXEC_NUM') if r.get('PERC_EXEC_NUM') is not None else -1
        c_exec = cur.get('PERC_EXEC_NUM') if cur.get('PERC_EXEC_NUM') is not None else -1
        if (r_exec, r.get('US EXECUTADO_NUM') or 0, r.get('VALOR EXEC_NUM') or 0) > (c_exec, cur.get('US EXECUTADO_NUM') or 0, cur.get('VALOR EXEC_NUM') or 0):
            best[key] = r
    return list(best.values())

def parse_filter_date(s):
    return date_value(s) if s else None


def apply_filters(rows, equipe='', area='', inicio='', fim=''):
    d_ini = parse_filter_date(inicio)
    d_fim = parse_filter_date(fim)
    out = []
    for r in rows:
        if equipe and str(r.get('EQUIPE') or '') != equipe:
            continue
        if area and str(r.get('AREA') or '') != area:
            continue
        d = r.get('DATA_OBJ')
        if d_ini and (d is None or d < d_ini):
            continue
        if d_fim and (d is None or d > d_fim):
            continue
        out.append(r)
    return out


def round2(n):
    return round(float(n or 0), 2)


def build_dashboard(params):
    rows_all = load_rows()
    equipe = params.get('equipe', [''])[0]
    area = params.get('area', [''])[0]
    inicio = params.get('inicio', [''])[0]
    fim = params.get('fim', [''])[0]

    all_equipes = sorted({str(r.get('EQUIPE')) for r in rows_all if r.get('EQUIPE')})
    all_areas = sorted({str(r.get('AREA')) for r in rows_all if r.get('AREA')})

    rows_all_filtered_raw = apply_filters(rows_all, equipe, area, inicio, fim)

    # Regra global dos gráficos/KPIs:
    # toda conta usa somente 1 registro por Nº NOTA, ignorando duplicações de AUX/repetições.
    rows_all_filtered = dedupe_by_nota(rows_all_filtered_raw)

    # Regra do painel:
    # - PROGRAMADO entra com as NOTAS únicas válidas da programação.
    # - EXECUTADO entra só quando há apontamento de execução preenchido.
    # - PERDA/MOTIVO entra só quando há execução apontada e ela é menor que 100%.
    rows_exec = [r for r in rows_all_filtered if r.get('EXEC_APONTADA')]

    by_date_map = defaultdict(lambda: {'US PLANEJADO': 0.0, 'US EXECUTADO': 0.0, 'VALOR PROG': 0.0, 'VALOR EXEC': 0.0, 'DATA_OBJ': None})
    for r in rows_all_filtered:
        label = r.get('DATA_LABEL') or 'Sem data'
        by_date_map[label]['US PLANEJADO'] += r['US PLANEJADO_NUM']
        by_date_map[label]['VALOR PROG'] += r['VALOR PROG_NUM']
        by_date_map[label]['DATA_OBJ'] = r.get('DATA_OBJ')
    for r in rows_exec:
        label = r.get('DATA_LABEL') or 'Sem data'
        by_date_map[label]['US EXECUTADO'] += r['US EXECUTADO_NUM']
        by_date_map[label]['VALOR EXEC'] += r['VALOR EXEC_NUM']
        by_date_map[label]['DATA_OBJ'] = r.get('DATA_OBJ')
    series = []
    for label, v in by_date_map.items():
        series.append({'DATA_LABEL': label, 'DATA_ISO': v['DATA_OBJ'].isoformat() if v.get('DATA_OBJ') else None, 'US PLANEJADO': round2(v['US PLANEJADO']), 'US EXECUTADO': round2(v['US EXECUTADO']), 'VALOR PROG': round2(v['VALOR PROG']), 'VALOR EXEC': round2(v['VALOR EXEC']), '_d': v['DATA_OBJ']})
    series.sort(key=lambda x: (x['_d'] is None, x['_d'] or date.min))

    # Comparativo acumulado: mostra o total acumulado de programado x executado por data.
    acc_prog = 0.0
    acc_exec = 0.0
    acc_valor_prog = 0.0
    acc_valor_exec = 0.0
    for s in series:
        acc_prog += s['US PLANEJADO']
        acc_exec += s['US EXECUTADO']
        acc_valor_prog += s['VALOR PROG']
        acc_valor_exec += s['VALOR EXEC']
        s['US PLANEJADO'] = round2(acc_prog)
        s['US EXECUTADO'] = round2(acc_exec)
        s['VALOR PROG'] = round2(acc_valor_prog)
        s['VALOR EXEC'] = round2(acc_valor_exec)
        s.pop('_d', None)

    perdas_map = defaultdict(lambda: {'perda_us': 0.0, 'perda_valor': 0.0, 'obras': 0})
    for r in rows_exec:
        if (r.get('PERC_EXEC_NUM') or 0) < 1:
            m = r.get('MOTIVO_PERDA') or 'Outros'
            perdas_map[m]['perda_us'] += r['PERDA_US']
            perdas_map[m]['perda_valor'] += r['PERDA_VALOR']
            perdas_map[m]['obras'] += 1
    perdas = [{'MOTIVO_PERDA': k, 'perda_us': round2(v['perda_us']), 'perda_valor': round2(v['perda_valor']), 'obras': v['obras']} for k, v in perdas_map.items()]
    perdas.sort(key=lambda x: x['perda_us'], reverse=True)

    # Ranking por equipe: diferente dos outros gráficos, aqui a mesma NOTA pode aparecer
    # para mais de uma equipe. Quando isso acontece, o US executado da nota é dividido
    # igualmente entre as equipes distintas que participaram dela.
    equipe_map = defaultdict(lambda: {'planejado': 0.0, 'executado': 0.0, 'perda': 0.0, 'obras': 0.0})
    notas_equipes = defaultdict(list)
    for r in rows_all_filtered_raw:
        key = nota_key(r)
        if key:
            # Para o Ranking por equipe, participantes da nota incluem TODAS as equipes
            # que apareceram na nota, mesmo se a linha delas não tiver % EXEC preenchido.
            # Ex.: se só uma linha está 100%, mas Anderson/Nilson/Roberto/Marciano
            # participaram da mesma nota, divide igualmente entre os 4.
            notas_equipes[key].append(r)

    for key, nota_rows in notas_equipes.items():
        exec_rows = [r for r in nota_rows if r.get('EXEC_APONTADA')]
        if not exec_rows:
            continue

        equipes_nota = []
        seen_equipe = set()
        for r in nota_rows:
            e = str(r.get('EQUIPE') or 'Sem equipe')
            if e not in seen_equipe:
                seen_equipe.add(e)
                equipes_nota.append(e)
        if not equipes_nota:
            continue

        base = max(exec_rows, key=lambda r: (
            r.get('PERC_EXEC_NUM') if r.get('PERC_EXEC_NUM') is not None else -1,
            r.get('US EXECUTADO_NUM') or 0,
            r.get('VALOR EXEC_NUM') or 0
        ))
        divisor = len(equipes_nota)
        exec_share = (base.get('US EXECUTADO_NUM') or 0) / divisor
        perda_share = (base.get('PERDA_US') or 0) / divisor
        obra_share = 1 / divisor
        for e in equipes_nota:
            equipe_map[e]['executado'] += exec_share
            equipe_map[e]['perda'] += perda_share
            equipe_map[e]['obras'] += obra_share

    # Mantém o campo planejado preenchido sem interferir no ranking executado.
    for r in rows_all_filtered:
        e = str(r.get('EQUIPE') or 'Sem equipe')
        equipe_map[e]['planejado'] += r['US PLANEJADO_NUM']

    equipes = [{'EQUIPE': k, 'planejado': round2(v['planejado']), 'executado': round2(v['executado']), 'perda': round2(v['perda']), 'obras': round2(v['obras'])} for k, v in equipe_map.items()]
    equipes.sort(key=lambda x: x['executado'], reverse=True)

    prog = sum(r['US PLANEJADO_NUM'] for r in rows_all_filtered)
    execu = sum(r['US EXECUTADO_NUM'] for r in rows_exec)
    kpi = {
        'obras_total': len(rows_all_filtered),
        'obras_apontadas': len({str(r.get('Nº NOTA') or '').strip() for r in rows_exec if str(r.get('Nº NOTA') or '').strip() and (r.get('PERC_EXEC_NUM') or 0) >= 1}),
        'us_planejado': round2(prog),
        'us_executado': round2(execu),
        'perda_us': round2(sum(r['PERDA_US'] for r in rows_exec)),
        'valor_planejado': round2(sum(r['VALOR PROG_NUM'] for r in rows_all_filtered)),
        'valor_executado': round2(sum(r['VALOR EXEC_NUM'] for r in rows_exec)),
        'aderencia': round(execu / prog * 100, 1) if prog else 0
    }

    linhas = []
    for r in sorted(rows_exec, key=lambda x: x.get('DATA_OBJ') or date.min, reverse=True)[:100]:
        linhas.append({
            'DATA': r.get('DATA_ISO'),
            'EQUIPE': r.get('EQUIPE') or '',
            'Nº NOTA': r.get('Nº NOTA') or '',
            'AREA': r.get('AREA') or '',
            '% EXEC': r.get('% EXEC') if r.get('% EXEC') is not None else '',
            'US PLANEJADO': round2(r.get('US PLANEJADO_NUM')),
            'US EXECUTADO': round2(r.get('US EXECUTADO_NUM')),
            'PERDA_US': round2(r.get('PERDA_US')),
            'VALOR EXEC': round2(r.get('VALOR EXEC_NUM')),
            'MOTIVO_PERDA': r.get('MOTIVO_PERDA') or '',
            'OBS.': r.get('OBS.') or ''
        })

    return {
        'kpi': kpi,
        'series': series,
        'perdas': perdas,
        'equipes': equipes,
        'filtros': {'equipes': all_equipes, 'areas': all_areas},
        'linhas': linhas
    }


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == '/api/dashboard':
            try:
                data = build_dashboard(parse_qs(parsed.query))
                body = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                body = json.dumps({'erro': str(e)}, ensure_ascii=False).encode('utf-8')
                self.send_response(500)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(body)
            return
        if parsed.path == '/':
            self.path = '/index.html'
        return super().do_GET()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', '5052'))
    print(f'CAD 2026 rodando na porta {port}')
    ThreadingHTTPServer(('0.0.0.0', port), Handler).serve_forever()
